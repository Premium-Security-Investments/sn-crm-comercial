import importlib.util
import json
import tempfile
import unittest
from pathlib import Path

import openpyxl


MODULE_PATH = Path(__file__).resolve().parents[1] / "scripts" / "extract_siio_board_sources.py"


def load_module():
    spec = importlib.util.spec_from_file_location("extract_siio_board_sources", MODULE_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


class SiioBoardSourceExtractorTest(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.base = Path(self.tmp.name)

    def tearDown(self):
        self.tmp.cleanup()

    def make_pyg(self):
        path = self.base / "pyg.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparat"
        ws.append([None, "Concepto", "2026", "Participación", None, "2025", "Participación", None, "Variación absoluta", "Variación %"])
        ws.append([None, "INGRESOS", 1000, 1.0, None, 800, 1.0, None, 200, 0.25])
        ws.append([None, "UTILIDAD NETA", 120, 0.12, None, 80, 0.10, None, 40, 0.50])
        ws.append([None, "NO APROBADO", 999, 0.99, None, 1, 0.01, None, 998, 998])
        wb.save(path)
        return path

    def make_payroll(self):
        path = self.base / "payroll.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        headers = ["Cedula", "Descripcion"] + [f"C{i}" for i in range(3, 20)] + ["DEV.", "DED.", "TOTAL", None, "Codigo", "Nombre Empleados", "Cargo", "Area", "DEV.", None, None]
        ws.append(headers)
        ws.append(["111", "Persona Uno"] + [None] * 17 + [100, 10, 90, None, "A", "Persona Uno", "Analista", "Finanzas", None, None, None])
        ws.append(["222", "Persona Dos"] + [None] * 17 + [200, 20, 180, None, "B", "Persona Dos", "Analista", "Finanzas", None, None, None])
        ws.append(["333", "Persona Tres"] + [None] * 17 + [150, 15, 135, None, "C", "Persona Tres", "Coordinador", "Operaciones", None, None, None])
        wb.save(path)
        return path

    def test_extracts_only_approved_financial_concepts(self):
        mod = load_module()
        rows = mod.extract_financial_metrics(self.make_pyg(), "2026-04-01", "SRC-PYG-2026-04")
        self.assertEqual([row["concept"] for row in rows], ["INGRESOS", "UTILIDAD NETA"])
        self.assertEqual(rows[0]["value_current"], 1000)
        self.assertEqual(rows[0]["value_comparison"], 800)
        self.assertEqual(rows[0]["variation_abs"], 200)
        self.assertEqual(rows[0]["variation_pct"], 0.25)
        self.assertEqual(rows[0]["source_id"], "SRC-PYG-2026-04")

    def test_payroll_is_aggregated_by_area_without_identity_fields(self):
        mod = load_module()
        rows = mod.extract_payroll_aggregates(self.make_payroll(), "2026-06-01", "SRC-NOMINA-2026-06")
        self.assertEqual([row["area"] for row in rows], ["Finanzas", "Operaciones"])
        finance = rows[0]
        self.assertEqual(finance["total_people"], 2)
        self.assertEqual(finance["total_accrued"], 300)
        self.assertEqual(finance["total_deductions"], 30)
        self.assertEqual(finance["net_total"], 270)
        self.assertEqual(finance["visibility_level"], "junta_agregado")
        serialized = json.dumps(rows, ensure_ascii=False).lower()
        for forbidden in ["cedula", "persona uno", "persona dos", "nombre empleados", "cargo"]:
            self.assertNotIn(forbidden, serialized)

    def test_rejects_payroll_without_required_headers(self):
        mod = load_module()
        path = self.base / "invalid.xlsx"
        wb = openpyxl.Workbook()
        wb.active.append(["Nombre", "Salario"])
        wb.save(path)
        with self.assertRaisesRegex(ValueError, "required payroll columns"):
            mod.extract_payroll_aggregates(path, "2026-06-01", "SRC")


if __name__ == "__main__":
    unittest.main()
