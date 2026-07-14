#!/usr/bin/env python3
"""Extract privacy-safe SIIO executive metrics from official board workbooks.

This script never emits employee names, IDs, roles or individual compensation.
It reads local copies only; downloading from or writing to SharePoint is outside its scope.
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

APPROVED_FINANCIAL_CONCEPTS = (
    "INGRESOS",
    "COSTOS",
    "GASTOS",
    "UTILIDAD OPERACIONAL",
    "MARGEN OPERACIONAL",
    "NO OPERACIONAL",
    "IMPUESTOS",
    "UTILIDAD NETA",
    "MARGEN NETO",
)


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


def financial_category(concept: str) -> str:
    if "MARGEN" in concept:
        return "margen"
    if concept == "INGRESOS":
        return "ingresos"
    if concept in {"COSTOS", "GASTOS", "IMPUESTOS"}:
        return "costos_gastos"
    return "resultado"


def extract_financial_metrics(path: str | Path, period_month: str, source_id: str) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=str(path).lower().endswith(".xlsm"))
    if "Comparat" not in workbook.sheetnames:
        raise ValueError("required financial sheet 'Comparat' not found")
    sheet = workbook["Comparat"]
    approved = set(APPROVED_FINANCIAL_CONCEPTS)
    metrics: list[dict[str, Any]] = []
    for row in sheet.iter_rows(values_only=True):
        concept = str(row[1] or "").strip().upper() if len(row) > 1 else ""
        if concept not in approved:
            continue
        metrics.append({
            "period_month": period_month,
            "category": financial_category(concept),
            "concept": concept,
            "value_current": number(row[2] if len(row) > 2 else None),
            "value_comparison": number(row[5] if len(row) > 5 else None),
            "variation_abs": number(row[8] if len(row) > 8 else None),
            "variation_pct": number(row[9] if len(row) > 9 else None),
            "source_id": source_id,
            "validated_by": None,
            "notes": "Extraído de la hoja Comparat; pendiente validación humana antes de publicación.",
        })
    return metrics


def normalized_header(value: Any) -> str:
    return str(value or "").strip().upper()


def payroll_columns(headers: tuple[Any, ...]) -> dict[str, int]:
    normalized = [normalized_header(value) for value in headers]
    try:
        area = normalized.index("AREA")
        accrued = normalized.index("DEV.")
        deductions = normalized.index("DED.")
        net = normalized.index("TOTAL")
    except ValueError as error:
        raise ValueError("required payroll columns not found: Area, DEV., DED., TOTAL") from error
    return {"area": area, "accrued": accrued, "deductions": deductions, "net": net}


def extract_payroll_aggregates(path: str | Path, period_month: str, source_id: str) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    aggregates: dict[str, dict[str, float]] = defaultdict(lambda: {"total_people": 0, "total_accrued": 0.0, "total_deductions": 0.0, "net_total": 0.0})
    found_sheet = False
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            continue
        try:
            columns = payroll_columns(headers)
        except ValueError:
            continue
        found_sheet = True
        for row in rows:
            area = str(row[columns["area"]] or "").strip() if len(row) > columns["area"] else ""
            if not area or area.lower() in {"total", "subtotal"}:
                continue
            item = aggregates[area]
            item["total_people"] += 1
            item["total_accrued"] += number(row[columns["accrued"]] if len(row) > columns["accrued"] else None)
            item["total_deductions"] += number(row[columns["deductions"]] if len(row) > columns["deductions"] else None)
            item["net_total"] += number(row[columns["net"]] if len(row) > columns["net"] else None)
    if not found_sheet:
        raise ValueError("required payroll columns not found: Area, DEV., DED., TOTAL")
    return [
        {
            "period_month": period_month,
            "area": area,
            "total_people": int(values["total_people"]),
            "total_accrued": round(values["total_accrued"], 2),
            "total_deductions": round(values["total_deductions"], 2),
            "net_total": round(values["net_total"], 2),
            "variation_abs": None,
            "alert": None,
            "source_id": source_id,
            "visibility_level": "junta_agregado",
        }
        for area, values in sorted(aggregates.items())
    ]


def build_snapshot(
    pyg_path: str | Path,
    payroll_path: str | Path,
    financial_period: str,
    payroll_period: str,
    financial_source_id: str,
    payroll_source_id: str,
) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "financial_period": financial_period,
        "payroll_period": payroll_period,
        "financial_metrics": extract_financial_metrics(pyg_path, financial_period, financial_source_id),
        "payroll_aggregates": extract_payroll_aggregates(payroll_path, payroll_period, payroll_source_id),
        "privacy": {
            "contains_personal_identifiers": False,
            "contains_individual_compensation": False,
            "payroll_visibility": "junta_agregado",
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pyg", required=True, type=Path)
    parser.add_argument("--payroll", required=True, type=Path)
    parser.add_argument("--financial-period", required=True)
    parser.add_argument("--payroll-period", required=True)
    parser.add_argument("--financial-source-id", default="SRC-PYG")
    parser.add_argument("--payroll-source-id", default="SRC-NOMINA")
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    snapshot = build_snapshot(args.pyg, args.payroll, args.financial_period, args.payroll_period, args.financial_source_id, args.payroll_source_id)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(args.output),
        "financial_metrics": len(snapshot["financial_metrics"]),
        "payroll_areas": len(snapshot["payroll_aggregates"]),
        "privacy": snapshot["privacy"],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
